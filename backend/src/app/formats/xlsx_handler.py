from pathlib import Path

from openpyxl import load_workbook

from app.formats.base import FormatHandler, register
from app.models.job import TextSegment


@register([".xlsx"])
class XlsxHandler(FormatHandler):
    def extract_texts(self, file_path: Path) -> list[TextSegment]:
        wb = load_workbook(str(file_path))
        segments = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        segments.append(TextSegment(
                            id=f"{sheet_name}:{cell.coordinate}",
                            text=cell.value,
                            context=f"Sheet: {sheet_name}, Cell: {cell.coordinate}",
                            metadata={"sheet": sheet_name, "coordinate": cell.coordinate},
                        ))
        return segments

    def replace_texts(self, file_path: Path, translated: list[TextSegment], output_path: Path) -> None:
        wb = load_workbook(str(file_path))
        lookup = {seg.id: seg.text for seg in translated}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    key = f"{sheet_name}:{cell.coordinate}"
                    if key in lookup:
                        cell.value = lookup[key]

        wb.save(str(output_path))
